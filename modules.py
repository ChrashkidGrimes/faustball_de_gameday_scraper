from __future__ import print_function
from datetime import datetime, timedelta, date
from cal_setup import get_calendar_service
from dateutil.relativedelta import relativedelta
import json
import requests
# scrape the webpage using headless webdriver
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import time
import openpyxl

# custom modules (place in same folder as this script!)
import settings

def getKey(val):
    for key, value in settings.leagues.items():
         if val == value:
             return key
    return "key doesn't exist"

def createEvent(summary, location, description, dateObjectStart, contestCat1Id):
   # creates one hour event tomorrow 10 AM IST
    service = get_calendar_service()
    event = {
    'summary': summary,
    'location': location,
    'description': description,
    'start': {
        'dateTime': dateObjectStart.strftime("%Y-%m-%dT%H:%M:%S"),
        'timeZone': 'Europe/Berlin',
    },
    'end': {
        'dateTime': (dateObjectStart + timedelta(hours=5)).strftime("%Y-%m-%dT%H:%M:%S"),
        'timeZone': 'Europe/Berlin',
    },
    }
    print(event)
    #event = service.events().insert(calendarId=settings.googleCalendarId[contestCat1Id], body=event).execute()
    print('Event created: %s' % (event.get('htmlLink')))

def createLongtermEvent(summary, dateObjectStart, category):
    service = get_calendar_service()
    event = {
    'summary': summary,
    'start': {
        'date': dateObjectStart,
        'timeZone': 'Europe/Berlin',
    },
    'end': {
        'date': dateObjectStart,
        'timeZone': 'Europe/Berlin',
    },
    }
    print(event)
    #event = service.events().insert(calendarId=settings.googleCalendarId[category], body=event).execute()
    print('Event created: %s' % (event.get('htmlLink')))

def getFrameDates():
    options = Options()
    options.headless = True
    # use firefox as a headless browser
    driver = webdriver.Firefox(options=options)
    # get web page
    driver.get(settings.urlpage)
    # sleep for 2s
    time.sleep(2)

    mySelect = driver.find_element_by_partial_link_text("Langzeitplanung")

    url = mySelect.get_attribute('href')
    r = requests.get(url, allow_redirects=True)

    open('Terminübersicht_Langzeitplanung.xlsx', 'wb').write(r.content)

    excel_file = openpyxl.load_workbook('Terminübersicht_Langzeitplanung.xlsx', data_only=True)
    # TODO: [FISTCRAWL-4] select excel sheet via season['name'] from getSeason()
    excel_sheet = excel_file['Feld 2021']

    # rows sind um 1 nach unten verschoben von der zählweise

    # inhalte einzelner zellen in excel_sheet -> _cells
    for league in settings.our_leagues:
        summary = f"{getKey(league)} - Spieltag"
        if "Männer" in getKey(league):
            category = 1
        elif "Frauen" in getKey(league):
            category = 2
        elif "M35" in getKey(league):
            category = 4
        else:
            category = 3


        for col in excel_sheet.iter_cols(min_row=7, min_col=league, max_col=league, max_row=70):
            for cell in col:
                if cell.value in ("X","1RR","2RR","3RR","E","AS","DM","SDM","WM",):
                    #print (excel_sheet.cell(row=cell.row, column=1).value.strftime("%d.%m.%Y"))
                    dateObjectStart = excel_sheet.cell(row=cell.row, column=1).value.strftime("%Y-%m-%d")
                    type(dateObjectStart)
                    # mit date kann ich schon einen Kalendereintrag machen
                    createLongtermEvent(summary, dateObjectStart, category)
    # close driver 
    driver.quit()

#TODO: [FISTCRAWL-2] get clubID from faustball.de with a search of club name which will be defined in settings.py

def getSeason():
    seasonIDs = []
    responseSeasons = requests.get(f"{settings.api}/core/seasons")
    allSeasons = json.loads(responseSeasons.text)
    # current date = actuall season plus 1
    for season in allSeasons:
        if datetime.strptime(season['startDay'], "%Y-%m-%d").date() <= (date.today() + relativedelta(months=6)) <= datetime.strptime(season['endDay'], "%Y-%m-%d").date():
            #print(season['name'])
            seasonIDs.append(season['id'])
    return seasonIDs

def getMatches(teamId, season):
    responseMatches = requests.get(f"{settings.api}/team/{teamId}/season/{season}/matches")
    spiele = json.loads(responseMatches.text)
    for values in spiele:
        # league
        responseContest = requests.get(f"{settings.api}/contest/{values['contestId']}")
        contest = json.loads(responseContest.text)
        # region
        responseRegion = requests.get(f"{settings.api}/region/{contest['regionId']}")
        region = json.loads(responseRegion.text)
        print(f"{contest['infoText']}")

        for locations in values['locations']:
            # host
            responseHost = requests.get(f"{settings.api}/club/{locations['ausrichterId']}")
            host = json.loads(responseHost.text)
            #print(host['longName'])
            # summary for event
            # contest['name'] -> Liga "1.Bundesliga", region['name'] -> Region "Süd", contest['cat1Id'] -> Kategorie "1" -> Männer
            
            if "Waldrennach" in host['longName']:
                summary = "Heimspieltag "
            else:
                summary="Auswärtsspieltag "
            # date and time of beginn matchday
            summary += str(f"{contest['name']} {region['name']} {settings.teamCategory[contest['cat1Id']]}")
            dateObjectStart = datetime.strptime(locations['start'], '%d.%m.%Y %H:%M')
            # venue
            responseVenue = requests.get(f"{settings.api}/venue/{locations['venueId']}")
            venue = json.loads(responseVenue.text)
            if locations['venueId'] == 0:
                VenueAddress = f"{host['longName']} - es ist keine Sportstätte hinterlegt"
                location = f"{venue['name']}\n{VenueAddress}"
            else:
                responseVenueAddress = requests.get(f"{settings.api}/venue/{locations['venueId']}/address")
                VenueAddress = json.loads(responseVenueAddress.text)
                location = f"{venue['name']}\n{VenueAddress['street']} {VenueAddress['houseNo']}, {VenueAddress['zip']} {VenueAddress['city']}"
            # games
            games = []
            for teams in locations['matches']:
                games.append(str(f"Durchgang: {teams['durchgang']}: {teams['teamA']} vs. {teams['teamB']}"))
            games.append(str(f"\n\n{contest['infoText']}"))
            schickeAusgabe = "\n".join(games)
            description = f"{schickeAusgabe}"
            createEvent(summary, location, description, dateObjectStart, contest['cat1Id'])